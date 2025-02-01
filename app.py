from flask import Flask, request, jsonify
from flask_cors import CORS  # Import the CORS library
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

app = Flask(__name__)

# Enable CORS for all routes
CORS(app)

@app.route('/save-defect', methods=['POST'])
def save_defect():
    data = request.json
    date_str = datetime.now().strftime("%Y-%m-%d")  # Format date as "YYYY-MM-DD"
    file_name = f"{date_str}_defect_record.xlsx"

    # Check if file exists, if not, create it and add a header row
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Date", "Time", "Line 1", "Line 2", "Line 3"])  # Header row

    # Append the defect data row
    sheet.append([data['date'], data['time'], data['line1'], data['line2'], data['line3']])
    workbook.save(file_name)

    return jsonify({"status": "success", "message": "Defect recorded successfully."})

if __name__ == "__main__":
    app.run(debug=True)
